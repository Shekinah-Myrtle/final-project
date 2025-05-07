import pandas as pd
import numpy as np
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook



# Path to Excel file
file_path = 'CTPROJECTDB.xlsx'

# Read all sheets;
all_sheets = pd.read_excel(file_path, sheet_name=None)

for sheet_name, df in all_sheets.items():
    print(f"Sheet: {sheet_name!r} — {df.shape[0]} rows x {df.shape[1]} columns")
input_file = 'CTPROJECTDB.xlsx'

# Load all sheets into a dictionary
all_sheets = pd.read_excel(input_file, sheet_name=None)

# Define material sheets 
material_sheets = [
    'Aggregates_Sand', 'Aluminium', 'Asphalt', 'Bitumen', 'Cement_and_Mortar',
    'Ceramic', 'Clay_Bricks', 'Concrete', 'Glass', 'Insulation', 'Paint',
    'Plaster', 'Rubber', 'Steel', 'Timber', 'Vinyl'
]

# Create a new Excel file 
output_file = 'Material_Data_Only.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name in material_sheets:
        # Check if the sheet exists in the original Excel file
        if sheet_name in all_sheets:
            # Extract the material data for the current sheet
            df = all_sheets[sheet_name]

            # Optionally, clean up the data (e.g., drop empty rows/columns)
            df_cleaned = df.dropna(how='all').dropna(axis=1, how='all')

            # Write the cleaned data to the new Excel file
            df_cleaned.to_excel(writer, sheet_name=sheet_name[:31], index=False)

print(f" Data from material sheets has been written to {output_file}")

# Path to your input Excel file
input_file = 'Material_Data_Only.xlsx'

# Define rows to drop for each material
rows_to_drop = {
    'Aggregates_Sand': list(range(0, 28)) + list(range(39, 49)) + list(range(85, 104)),
    'Aluminium': list(range(0, 30)) + list(range(59, 69)) + list(range(105, 124)) + list(range(132, 142)) + list(range(147, 299)) + list(range(384, 1931)),
    'Asphalt': list(range(0, 30)) + list(range(52, 62)) + list(range(66, 80)) + list(range(89, 117)) + list(range(122, 135)) + list(range(137, 138)) + list(range(144, 291)) + list(range(314, 1923)),
    'Bitumen': list(range(0, 28)) + list(range(32, 42)) + list(range(48, 61)) + list(range(78, 97)) + list(range(104, 118)) + list(range(121, 270)) + list(range(288, 1904)),
    'Cement_and_Mortar': list(range(0, 28)) + list(range(87, 97)) + list(range(105, 116)) + list(range(125, 152)) + list(range(161, 172)) + list(range(186, 325)) + list(range(398, 1957)),
    'Ceramic': list(range(0, 28)) + list(range(32, 42)) + list(range(47, 61)) + list(range(70, 99)) + list(range(104, 116)) + list(range(122, 134)) + list(range(144, 287)) + list(range(441, 1919)),
    'Clay_Bricks': list(range(0, 28)) + list(range(0, 60)) + list(range(66, 78)) + list(range(88, 115)) + list(range(122, 135)) + list(range(149, 288)) + list(range(307, 1920)),
    'Concrete': list(range(0, 30)) + list(range(310, 319)) + list(range(326, 338)) + list(range(345, 375)) + list(range(382, 395)) + list(range(625, 2180)),
    'Glass': list(range(0, 28)) + list(range(95, 104)) + list(range(113, 122)) + list(range(133, 160)) + list(range(169, 178)) + list(range(199, 333)) + list(range(525, 1965)),
    'Insulation': list(range(0, 28)) + list(range(34, 43)) + list(range(51, 62)) + list(range(72, 100)) + list(range(108, 118)) + list(range(126, 138)) + list(range(181, 292)) + list(range(1259, 1923)),
    'Paint': list(range(0, 28)) + list(range(30, 39)) + list(range(43, 57)) + list(range(68, 97)) + list(range(100, 114)) + list(range(118, 134)) + list(range(874, 1919)),
    'Plaster': list(range(0, 28)) + list(range(32, 41)) + list(range(47, 60)) + list(range(70, 98)) + list(range(104, 116)) + list(range(122, 136)) + list(range(161, 289)) + list(range(860, 1922)),
    'Rubber': list(range(0, 28)) + list(range(30, 39)) + list(range(43, 58)) + list(range(68, 96)) + list(range(100, 114)) + list(range(118, 134)) + list(range(140, 287)) + list(range(351, 1919)),
    'Steel': list(range(0, 28)) + list(range(48, 76)) + list(range(134, 143)) + list(range(151, 162)) + list(range(171, 199)) + list(range(2067, 219)) + list(range(224, 372)) + list(range(525, 2003)),
    'Timber': list(range(0, 28)) + list(range(74, 83)) + list(range(112, 139)) + list(range(153, 159)) + list(range(205, 312)) + list(range(526, 1944)),
    'Vinyl': list(range(0, 28)) + list(range(31, 40)) + list(range(45, 58)) + list(range(69, 97)) + list(range(102, 115)) + list(range(120, 135)) + list(range(138, 288)) + list(range(568, 1920))
}

# Load all sheets into a dictionary
all_sheets = pd.read_excel(input_file, sheet_name=None)

# Create a new Excel file to save updated material data
output_file = 'Updated_Material_Data.xlsx'

# Delete the file if it exists
if os.path.exists(output_file):
    os.remove(output_file)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name, df in all_sheets.items():
        if sheet_name in rows_to_drop:
            rows = [i for i in rows_to_drop[sheet_name] if i < len(df)]
            df = df.drop(df.index[rows]).reset_index(drop=True)
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

print(f" Updated material data has been written to {output_file}")

# Load the workbook
wb = load_workbook('Updated_Material_Data.xlsx')

# Iterate over each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}")
    
    # Delete columns A to D (1 to 4)
    for col in range(4, 0, -1):
        ws.delete_cols(col)
        print(f"Deleted column {col}")
    
    # Find the maximum number of columns after deleting A to D
    max_col = ws.max_column
    
    # Delete columns O to Z (15 to max_col)
    for col in range(max_col, 14, -1):
        ws.delete_cols(col)
        print(f"Deleted column {col}")

# Save the workbook as a new file in the current working directory
current_dir = os.getcwd()
new_file_name = 'Updated_Material_Data_New.xlsx'
wb.save(os.path.join(current_dir, new_file_name))

print(f"File saved as {new_file_name} in {current_dir}")

# Load the workbook
wb = load_workbook('Updated_Material_Data_New.xlsx')

# Iterate over each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}")
    
    # Delete columns K to Z (11 and above)
    while ws.max_column > 10:
        ws.delete_cols(11)

# Save the workbook to a new file
new_file_name = 'Updated_Material_Data_Final.xlsx'
if os.path.exists(new_file_name):
    os.remove(new_file_name)
wb.save(new_file_name)

print(f"Columns K to Z deleted and saved to {new_file_name}")

# Load the workbook
wb = load_workbook('Updated_Material_Data_Final.xlsx')

# Define the ranges to clear for each material
ranges_to_clear = {
    'Aggregates_Sand': [(1, 12, 'C', 'G')],
    'Aluminium': [(1, 7, 'C', 'G'), (9, 14, 'C', 'G'), (16, 30, 'C', 'G')],
    'Asphalt': [(1, 11, 'C', 'G'), (14, 23, 'C', 'G')],
    'Bitumen': [(1, 4, 'C', 'G')],
    'Cement_and_Mortar': [(1, 5, 'C', 'G'), (7, 8, 'C', 'G'), (13, 16, 'C', 'G'), (18, 21, 'C', 'G'), (23, 24, 'C', 'G'), (26, 28, 'C', 'G'), (30, 31, 'C', 'G'), (33, 34, 'C', 'G'), (36, 60, 'C', 'G')],
    'Ceramic': [(1, 5, 'C', 'G')],
    'Concrete': [(1, 57, 'C', 'G'), (58, 86, 'C', 'G'), (88, 144, 'E', 'E'), (146, 195, 'A', 'F'), (191, 195, 'F', 'F'), (197, 200, 'F', 'F'), (201, 204, 'A', 'G'), (206, 281, 'E', 'E')],
    'Glass': [(1, 68, 'C', 'G')],
    'Insulation': [(1, 7, 'C', 'G')],
    'Paint': [(1, 4, 'C', 'G')],
    'Plaster': [(1, 5, 'C', 'G')],
    'Rubber': [(1, 3, 'C', 'G')],
    'Steel': [(1, 21, 'C', 'G')],
    'Timber': [(1, 47, 'C', 'G')],
    'Vinyl': [(1, 4, 'C', 'G')]
}

# Function to convert column letter to column index
def column_to_index(column):
    index = 0
    for char in column:
        index = index * 26 + ord(char.upper()) - ord('A') + 1
    return index

# Iterate over each sheet
for sheet_name in wb.sheetnames:
    if sheet_name in ranges_to_clear:
        ws = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")
        
        # Clear comments in specified columns
        for start_row, end_row, start_col, end_col in ranges_to_clear[sheet_name]:
            start_col_index = column_to_index(start_col)
            end_col_index = column_to_index(end_col)
            for row in range(start_row, end_row + 1):
                for col in range(start_col_index, end_col_index + 1):
                    ws.cell(row=row, column=col).value = ""
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.delete_rows(1)

# Shift data from columns H and I to columns C and D
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=3).value = ws.cell(row=row, column=8).value
        ws.cell(row=row, column=4).value = ws.cell(row=row, column=9).value
        ws.cell(row=row, column=8).value = None
        ws.cell(row=row, column=9).value = None

# Save the workbook
new_file_name = 'Updated_Material_Data_Final_Cleaned.xlsx'
if os.path.exists(new_file_name):
    os.remove(new_file_name)
wb.save(new_file_name)

print(f"saved to {new_file_name}")


# Load the workbook
wb = load_workbook('Updated_Material_Data_Final_Cleaned.xlsx')

# Define the ranges to copy for each material
ranges_to_copy = {
    'Aggregates_Sand': [(2, 11, 'A', 'D')],
    'Aluminium': [(2, 29, 'A', 'D')],
    'Asphalt': [(2, 22, 'A', 'D')],
    'Bitumen': [(2, 4, 'A', 'D')],
    'Cement_and_Mortar': [(3, 59, 'A', 'D')],
    'Ceramic': [(2, 4, 'A', 'D')],
    'Concrete': [(5, 143, 'A', 'D')],
    'Glass': [(2, 67, 'A', 'D')],
    'Insulation': [(2, 6, 'A', 'D')],
    'Paint': [(2, 2, 'A', 'D')],
    'Plaster': [(2, 4, 'A', 'D')],
    'Rubber': [(2, 2, 'A', 'D')],
    'Steel': [(5, 20, 'A', 'D')],
    'Timber': [(4, 46, 'A', 'D')],
    'Vinyl': [(2, 3, 'A', 'D')]
}

# Function to convert column letter to column index
def column_to_index(column):
    index = 0
    for char in column:
        index = index * 26 + ord(char.upper()) - ord('A') + 1
    return index

# Create a new workbook
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = 'Material DQI'

# Set the header
new_ws.cell(row=1, column=1).value = 'Materials'
new_ws.cell(row=1, column=2).value = 'Embodied Carbon - kgCO2e/kg'
new_ws.cell(row=1, column=3).value = 'DQI Score'
new_ws.cell(row=1, column=4).value = 'DQI Version'

row_index = 2
for sheet_name in ranges_to_copy.keys():
    ws = wb[sheet_name]
    for start_row, end_row, start_col, end_col in ranges_to_copy[sheet_name]:
        start_col_index = column_to_index(start_col)
        end_col_index = column_to_index(end_col)
        for row in range(start_row, end_row + 1):
            new_ws.cell(row=row_index, column=1).value = ws.cell(row=row, column=1).value
            new_ws.cell(row=row_index, column=2).value = ws.cell(row=row, column=2).value
            new_ws.cell(row=row_index, column=3).value = ws.cell(row=row, column=3).value
            new_ws.cell(row=row_index, column=4).value = ws.cell(row=row, column=4).value
            row_index += 1

# Save the new workbook
new_file_name = 'FinalDataSet.xlsx'
if os.path.exists(new_file_name):
    os.remove(new_file_name)
new_wb.save(new_file_name)
 
print(f"saved to {new_file_name}")


# Load the workbook
wb1 = load_workbook('Updated_Material_Data_Final_Cleaned.xlsx')
wb2 = load_workbook('FinalDataSet.xlsx')

# Define the ranges to copy for each material
ranges_to_copy = {
    'Aggregates_Sand': (13, 20),
    'Aluminium': (31, 36),
    'Asphalt': (24, 26),
    'Bitumen': (6, 10),
    'Cement_and_Mortar': (61, 67),
    'Ceramic': (6, 9),
    'Clay_Bricks': (2, 6),
    'Concrete': (283, 286),
    'Glass': (70, 76),
    'Insulation': (9, 14),
    'Paint': (5, 6),
    'Plaster': (7, 10),
    'Rubber': (5, 6),
    'Steel': (81, 86),
    'Timber': (49, 60),
    'Vinyl': (6, 8)
}

# Create a new sheet in the FinalDataSet.xlsx file
if 'Material Data' in wb2.sheetnames:
    wb2.remove(wb2['Material Data'])
new_ws = wb2.create_sheet('Material Data')

# Set the header
new_ws.cell(row=1, column=1).value = 'Main Material'
new_ws.cell(row=1, column=2).value = 'Sample Size'
new_ws.cell(row=1, column=3).value = 'Max Embodied Carbon (kg CO2e/kg)'
new_ws.cell(row=1, column=4).value = 'Standard Deviation Embodied Carbon (kg CO2e/kg)'
new_ws.cell(row=1, column=5).value = 'Variance Based on a Sample (Excel Var Function) Embodied Carbon (kg CO2e/kg)'

row_index = 2
for sheet_name, (start_row, end_row) in ranges_to_copy.items():
    ws = wb1[sheet_name]
    for row in range(start_row, end_row + 1):
        new_ws.cell(row=row_index, column=1).value = ws.cell(row=row, column=1).value
        new_ws.cell(row=row_index, column=2).value = ws.cell(row=row, column=2).value
        new_ws.cell(row=row_index, column=3).value = ws.cell(row=row, column=5).value
        new_ws.cell(row=row_index, column=4).value = ws.cell(row=row, column=6).value
        new_ws.cell(row=row_index, column=5).value = ws.cell(row=row, column=7).value
        row_index += 1

# Save the workbook
wb2.save('FinalDataSet.xlsx')

print(f"Material data created and saved to FinalDataSet.xlsx")


# Load the data
df = pd.read_excel('FinalDataSet.xlsx', sheet_name='Material DQI')

# Rename material column for specific rows
material_renames = {
    67: 'Portland Slag Cement',
    68: 'Portland Slag Cement',
    70: 'Portland Pozzolana Cement',
    71: 'Portland Pozzolana Cement',
    73: 'Portland Fly Ash Cement',
    74: 'Portland Fly Ash Cement',
    75: 'Portland Fly Ash Cement',
    76: 'Portland Fly Ash Cement',
    78: 'Portland Limestone Cement',
    79: 'Portland Limestone Cement',
    80: 'Portland Limestone Cement',
    81: 'Portland Limestone Cement',
    83: 'Portland Composite Cement',
    84: 'Portland Composite Cement',
    86: 'Blast Furnace Cement',
    87: 'Blast Furnace Cement',
    88: 'Blast Furnace Cement',
    90: 'Pozzolanoc Cement',
    91: 'Pozzolanoc Cement',
    93: 'Composite Cement',
    94: 'Composite Cement'
}

for index, new_name in material_renames.items():
    df.loc[index-1, 'Materials'] = new_name

# Remove specific rows
rows_to_remove = [17, 18, 24, 25, 26, 49, 50, 51, 66, 69, 72, 77, 82, 85, 89, 92, 95, 96, 104, 112, 113, 202, 203, 204, 270, 271, 273, 284, 294, 301, 308, 315, 327, 375, 376, 377]

df = df.drop([row-1 for row in rows_to_remove], axis=0).reset_index(drop=True)

# Save the updated DataFrame to the Excel file
with pd.ExcelWriter('FinalDataSet.xlsx', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name='Material DQI', index=False)

# Load the workbook
material_wb = load_workbook('Updated_Material_Data_Final_Cleaned.xlsx')
wb = load_workbook('FinalDataSet.xlsx')

# Create a new sheet
try:
    new_ws = wb['Material_Data1']
    wb.remove(new_ws)
    new_ws = wb.create_sheet('Material_Data1')
except KeyError:
    new_ws = wb.create_sheet('Material_Data1')

# Set the header
headers = ['Main Material', 'Sample Size', 'DQI Sample Size (Max 10)', 'DQI Total - % (Max 100%)', 
           'DQI Temporal (Max 5)', 'DQI Geographic (Max 5)', 'DQI Transparency (Max 5)']

for col, header in enumerate(headers, start=1):
    new_ws.cell(row=1, column=col).value = header
    
# Define the row ranges for each material sheet
ranges_to_copy = {
    'Aggregates_Sand': (50, 56),
    'Aluminium': (68, 73),
    'Asphalt': (38, 40),
    'Bitumen': (30, 34),
    'Cement_and_Mortar': (79, 85),
    'Clay_Bricks': (19, 23),
    'Concrete': (297, 300),
    'Glass': (90, 96),
    'Insulation': None,  # Add row range if needed
    'Paint': None,  # Add row range if needed
    'Plaster': None,  # Add row range if needed
    'Rubber': None,  # Add row range if needed
    'Steel': (98, 103),
    'Timber': (78, 89),
    'Vinyl': None  # Add row range if needed
}

row_index = 2
for sheet_name, row_range in ranges_to_copy.items():
    if row_range is not None:
        ws = material_wb[sheet_name]
        for row in range(row_range[0], row_range[1] + 1):
            for col in range(1, 8):
                new_ws.cell(row=row_index, column=col).value = ws.cell(row=row, column=col).value
            row_index += 1

# Save the workbook
wb.save('FinalDataSet.xlsx')



# Load the workbook
material_wb = load_workbook('Updated_Material_Data_Final_Cleaned.xlsx')
wb = load_workbook('FinalDataSet.xlsx')

# Create a new sheet
try:
    new_ws = wb['Material_Data2']
    wb.remove(new_ws)
    new_ws = wb.create_sheet('Material_Data2')
except KeyError:
    new_ws = wb.create_sheet('Material_Data2')

# Set the header
headers = ['Material', 'Density (kg m-3)', 'Specific heat (J kg-1 K-1)', 'Thermal Diffusivity (M^2 S-1)']

for col, header in enumerate(headers, start=1):
    new_ws.cell(row=1, column=col).value = header

# Define the row ranges for each material sheet
ranges_to_copy = {
    'Aggregates_Sand': (70, 72),
    'Aluminium': (78, 78),
    'Asphalt': (44, 46),
    'Bitumen': (36, 37),
    'Cement_and_Mortar': (88, 99),
    'Ceramic': (32, 39),
    'Clay_Bricks': (25, 37),
    'Concrete': (306, 428),
    'Glass': (101, 117),
    'Insulation': (44, 83),
    'Plaster': (35, 57),
    'Rubber': (27, 30),
    'Steel': (118, 120),
    'Timber': (92, 135),
    'Vinyl': (32, 32)
}

row_index = 2
for sheet_name, row_range in ranges_to_copy.items():
    ws = material_wb[sheet_name]
    for row in range(row_range[0], row_range[1] + 1):
        new_ws.cell(row=row_index, column=1).value = ws.cell(row=row, column=1).value
        new_ws.cell(row=row_index, column=2).value = ws.cell(row=row, column=5).value
        new_ws.cell(row=row_index, column=3).value = ws.cell(row=row, column=6).value
        new_ws.cell(row=row_index, column=4).value = ws.cell(row=row, column=7).value
        row_index += 1

# Save the workbook
wb.save('FinalDataSet.xlsx')

# Load the workbook
wb = load_workbook('FinalDataSet.xlsx')

# Get the Material_Data_2 sheet
ws = wb['Material_Data2']

# Delete rows with empty cells in column A
rows_to_delete = []
for row in range(1, ws.max_row + 1):
    if ws.cell(row=row, column=1).value is None:
        rows_to_delete.append(row)

# Delete rows in reverse order to avoid index issues
for row in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row)

# Get the Material DQI sheet
ws = wb['Material DQI']

# Delete rows
rows_to_delete = [116, 77, 75, 72, 70, 66, 62, 60, 58, 44]
for row in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row)



# Save the workbook
wb.save('FinalDataSet.xlsx')

#PLOTTING

# Load the Excel file
df = pd.read_excel('FinalDataSet.xlsx', sheet_name='Material DQI')

def extract_material_group(name):
    name = str(name).lower()
    if 'steel' in name:
        return 'Steel'
    elif 'concrete' in name:
        return 'Concrete'
    elif 'timber' in name:
        return 'Timber'
    elif 'glass' in name:
        return 'Glass'
    elif 'aluminum' in name or 'aluminium' in name:
        return 'Aluminum'
    elif 'asphalt' in name:
        return 'Asphalt'
    elif 'bitumen' in name:
        return 'Bitumen'
    elif 'cement' in name:
        return 'Cement'
    elif 'aggregate' in name:
        return 'Aggregate'
    elif 'clay' in name:
        return 'Clay'
    elif 'paint' in name:
        return 'Paint'
    elif 'vinyl' in name:
        return 'Vinyl'
    elif 'insulation' in name:
        return 'Insulation'
    elif 'rubber' in name:
        return 'Rubber'
    elif 'plaster' in name:
        return 'Plaster'
    else:
        return 'Other'

# Apply the function to the 'Material' column
df['Material Group'] = df['Materials'].apply(extract_material_group)

# Convert 'Embodied Carbon - kgCO2e/kg' column to numeric values
df['Embodied Carbon - kgCO2e/kg'] = pd.to_numeric(df['Embodied Carbon - kgCO2e/kg'], errors='coerce')

# Drop rows with non-numeric values
df = df.dropna(subset=['Embodied Carbon - kgCO2e/kg'])

# Group by material group and calculate mean embodied carbon
grouped_df = df.groupby('Material Group')['Embodied Carbon - kgCO2e/kg'].mean().reset_index()

# Create a bar chart of material group vs embodied carbon
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['Embodied Carbon - kgCO2e/kg'])
plt.xlabel('Material Group')
plt.ylabel('Embodied Carbon - kgCO2e/kg')
plt.title('Material Group vs Embodied Carbon')
plt.xticks(rotation=90)
plt.tight_layout()
plt.savefig('plot1.png',bbox_inches='tight')
plt.show()

#2nd plotting 
# Load the Excel file
df = pd.read_excel('FinalDataSet.xlsx', sheet_name='Material DQI')

def extract_material_group(name):
    name = str(name).lower()
    material_groups = {
        'steel': 'Steel',
        'concrete': 'Concrete',
        'timber': 'Timber',
        'glass': 'Glass',
        'aluminum': 'Aluminum',
        'aluminium': 'Aluminum',
        'asphalt': 'Asphalt',
        'bitumen': 'Bitumen',
        'cement': 'Cement',
        'aggregate': 'Aggregate',
        'clay': 'Clay',
        'paint': 'Paint',
        'vinyl': 'Vinyl',
        'insulation': 'Insulation',
        'rubber': 'Rubber',
        'plaster': 'Plaster'
    }
    for keyword, group in material_groups.items():
        if keyword in name:
            return group
    return 'Other'

# Apply the function to the 'Material' column
df['Material Group'] = df['Materials'].apply(extract_material_group)

# Convert 'DQI Score' column to numeric values
df['DQI Score'] = pd.to_numeric(df['DQI Score'], errors='coerce')

# Drop rows with non-numeric values
df = df.dropna(subset=['DQI Score'])

# Group by material group and calculate mean DQI Score
grouped_df = df.groupby('Material Group')['DQI Score'].mean().reset_index()

# Create a bar chart of material group vs DQI Score
plt.figure(figsize=(12, 8))
plt.bar(grouped_df['Material Group'], grouped_df['DQI Score'])
plt.xlabel('Material Group')
plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
plt.ylabel('DQI Score')
plt.title('Material Group vs DQI Score')
plt.xticks(rotation=90)

plt.tight_layout()
plt.savefig('plot2.png', bbox_inches='tight')
plt.show()

#3th plot
# Filter the data for Aggregate material
aggregate_df = df[df['Materials'].str.contains('aggregate', case=False, na=False)]

# Extract the aggregate type from each material name
aggregate_types = aggregate_df['Materials'].tolist()
abbreviations = []
for material in aggregate_types:
    abbreviation = ''.join(word[0].upper() for word in material.split())
    abbreviations.append(abbreviation)

# Create lists for metrics and values
metrics = ['Embodied Carbon - kgCO2e/kg','DQI Score']
x = np.arange(len(aggregate_df['Materials']))
width = 0.15

# Create a figure
fig, ax = plt.subplots(figsize=(25,10))  # Increased figure width

# Loop through each metric
for i, metric in enumerate(metrics):
    ax.bar(x + i * width, aggregate_df[metric], width, label=metric)

ax.set_xlabel('Aggregate Type', fontsize=14)
ax.set_ylabel('Values', fontsize=14)
ax.set_title('Metrics vs Values for Aggregate Material', fontsize=16)
ax.set_xticks(x + 0.5 * width * (len(metrics) - 1))
ax.set_xticklabels(abbreviations, rotation=45, ha='right', fontsize=12)
ax.legend(fontsize=12)

# Add a note with the abbreviations
note = '\n'.join(f'{abbreviation}: {material}' for abbreviation, material in zip(abbreviations, aggregate_types))
ax.annotate(note, xy=(1.05, 0.5), xycoords='axes fraction', ha='left', fontsize=12)

# Layout so plots do not overlap
fig.tight_layout(rect=[0,0,0.75,1])
plt.savefig('plot3.png',bbox_inches='tight')
plt.show()

#4TH PLOT
# Filter the data for Asphalt material
asphalt_df = df[df['Materials'].str.contains('asphalt', case=False, na=False)]

# Extract the asphalt type from each material name
asphalt_types = asphalt_df['Materials'].tolist()
abbreviations = []
for material in asphalt_types:
    abbreviation = ''.join(word[0].upper() for word in material.split())
    abbreviations.append(abbreviation)

# Create lists for metrics and values
metrics = ['Embodied Carbon - kgCO2e/kg','DQI Score']
x = np.arange(len(asphalt_df['Materials']))
width = 0.15

# Create a figure
fig, ax = plt.subplots(figsize=(25,10))

# Loop through each metric
for i, metric in enumerate(metrics):
    ax.bar(x + i * width, asphalt_df[metric], width, label=metric)

ax.set_xlabel('Asphalt Type', fontsize=14)
ax.set_ylabel('Values', fontsize=14)
ax.set_title('Metrics vs Values for Asphalt Material', fontsize=16)
ax.set_xticks(x + 0.5 * width * (len(metrics) - 1))
ax.set_xticklabels(abbreviations, rotation=45, ha='right', fontsize=12)
ax.legend(fontsize=12)

# Add a note with the abbreviations
note = '\n'.join(f'{abbreviation}: {material}' for abbreviation, material in zip(abbreviations, asphalt_types))
ax.annotate(note, xy=(1.05, 0.5), xycoords='axes fraction', ha='left', fontsize=12)

# Layout so plots do not overlap
fig.tight_layout(rect=[0,0,0.75,1])
plt.tight_layout()
plt.savefig('plot4.png', bbox_inches='tight')
plt.show()

#5TH PLOTTING
#Filter the data for Cement material
cement_df = df[df['Materials'].str.contains('cement', case=False, na=False)]

# Ensure the columns are numeric
cement_df['Embodied Carbon - kgCO2e/kg'] = pd.to_numeric(cement_df['Embodied Carbon - kgCO2e/kg'], errors='coerce')
cement_df['DQI Score'] = pd.to_numeric(cement_df['DQI Score'], errors='coerce')

# Drop rows with missing values
cement_df = cement_df.dropna(subset=['Embodied Carbon - kgCO2e/kg', 'DQI Score'])

# Extract the cement type from each material name
cement_types = cement_df['Materials'].tolist()
abbreviations = []
for material in cement_types:
    abbreviation = ''.join(word[0].upper() for word in material.split())
    abbreviations.append(abbreviation)

# Create lists for metrics and values
metrics = ['Embodied Carbon - kgCO2e/kg','DQI Score']
x = np.arange(len(cement_df['Materials']))
width = 0.15

# Create a figure
fig, ax = plt.subplots(figsize=(25,10))

# Loop through each metric
for i, metric in enumerate(metrics):
    ax.bar(x + i * width, cement_df[metric], width, label=metric)

ax.set_xlabel('Cement Type', fontsize=14)
ax.set_ylabel('Values', fontsize=14)
ax.set_title('Metrics vs Values for Cement Material', fontsize=16)
ax.set_xticks(x + 0.5 * width * (len(metrics) - 1))
ax.set_xticklabels(abbreviations, rotation=45, ha='right', fontsize=12)
ax.legend(fontsize=12)

# Add a note with the abbreviations
note = '\n'.join(f'{abbreviation}: {material}' for abbreviation, material in zip(abbreviations, cement_types))
ax.annotate(note, xy=(1.05, 0.5), xycoords='axes fraction', ha='left', fontsize=12)

# Layout so plots do not overlap
fig.tight_layout(rect=[0,0,0.75,1])

plt.tight_layout()
plt.savefig('plot5.png', bbox_inches='tight')
plt.show()

#6TH PLOTTING
# Filter the data for Ceramic material
ceramic_df = df[df['Materials'].str.contains('ceramic', case=False, na=False)]

# Ensure the columns are numeric
ceramic_df['Embodied Carbon - kgCO2e/kg'] = pd.to_numeric(ceramic_df['Embodied Carbon - kgCO2e/kg'], errors='coerce')
ceramic_df['DQI Score'] = pd.to_numeric(ceramic_df['DQI Score'], errors='coerce')

# Drop rows with missing values
ceramic_df = ceramic_df.dropna(subset=['Embodied Carbon - kgCO2e/kg', 'DQI Score'])

# Extract the ceramic type from each material name
ceramic_types = ceramic_df['Materials'].tolist()
abbreviations = []
for material in ceramic_types:
    abbreviation = ''.join(word[0].upper() for word in material.split())
    abbreviations.append(abbreviation)

# Create lists for metrics and values
metrics = ['Embodied Carbon - kgCO2e/kg','DQI Score']
x = np.arange(len(ceramic_df['Materials']))
width = 0.15

# Create a figure
fig, ax = plt.subplots(figsize=(25,10))

# Loop through each metric
for i, metric in enumerate(metrics):
    ax.bar(x + i * width, ceramic_df[metric], width, label=metric)
ax.set_xlabel('Ceramic Type', fontsize=14)
ax.set_ylabel('Values', fontsize=14)
ax.set_title('Metrics vs Values for Ceramic Material', fontsize=16)
ax.set_xticks(x + 0.5 * width * (len(metrics) - 1))
ax.set_xticklabels(abbreviations, rotation=45, ha='right', fontsize=12)
ax.legend(fontsize=12)

# Add a note with the abbreviations
note = '\n'.join(f'{abbreviation}: {material}' for abbreviation, material in zip(abbreviations, ceramic_types))
ax.annotate(note, xy=(1.05, 0.5), xycoords='axes fraction', ha='left', fontsize=12)

# Layout so plots do not overlap
fig.tight_layout(rect=[0,0,0.75,1])
plt.tight_layout()
plt.savefig('plot6.png',bbox_inches='tight')
plt.show()

#7TH PLOTTING
# Filter the data for Steel material
steel_df = df[df['Materials'].str.contains('steel', case=False, na=False)]

# Ensure the columns are numeric
steel_df['Embodied Carbon - kgCO2e/kg'] = pd.to_numeric(steel_df['Embodied Carbon - kgCO2e/kg'], errors='coerce')
steel_df['DQI Score'] = pd.to_numeric(steel_df['DQI Score'], errors='coerce')

# Drop rows with missing values
steel_df = steel_df.dropna(subset=['Embodied Carbon - kgCO2e/kg', 'DQI Score'])

# Extract the steel type from each material name
steel_types = steel_df['Materials'].tolist()
abbreviations = []
for material in steel_types:
    abbreviation = ''.join(word[0].upper() for word in material.split())
    abbreviations.append(abbreviation)

# Create lists for metrics and values
metrics = ['Embodied Carbon - kgCO2e/kg','DQI Score']
x = np.arange(len(steel_df['Materials']))
width = 0.15

# Create a figure
fig, ax = plt.subplots(figsize=(30,15))

# Loop through each metric
for i, metric in enumerate(metrics):
    ax.bar(x + i * width, steel_df[metric], width, label=metric)

ax.set_xlabel('Steel Type', fontsize=14)
ax.set_ylabel('Values', fontsize=14)
ax.set_title('Metrics vs Values for Steel Material', fontsize=16)
ax.set_xticks(x + 0.5 * width * (len(metrics) - 1))
ax.set_xticklabels(abbreviations, rotation=45, ha='right', fontsize=12)
ax.legend(fontsize=12)

# Add a note with the abbreviations
note = '\n'.join(f'{abbreviation}: {material}' for abbreviation, material in zip(abbreviations, steel_types))
ax.annotate(note, xy=(1.05, 0.5), xycoords='axes fraction', ha='left', fontsize=12)

# Layout so plots do not overlap
fig.tight_layout(rect=[0,0,0.75,1])
plt.tight_layout()
plt.savefig('plot7.png', bbox_inches='tight')
plt.show()

#8th plot
# Filter the data for Insulation material
insulation_df = df[df['Materials'].str.contains('insulation', case=False, na=False)]

# Ensure the columns are numeric
insulation_df['Embodied Carbon - kgCO2e/kg'] = pd.to_numeric(insulation_df['Embodied Carbon - kgCO2e/kg'], errors='coerce')
insulation_df['DQI Score'] = pd.to_numeric(insulation_df['DQI Score'], errors='coerce')

# Drop rows with missing values
insulation_df = insulation_df.dropna(subset=['Embodied Carbon - kgCO2e/kg', 'DQI Score'])

# Extract the insulation type from each material name
insulation_types = insulation_df['Materials'].tolist()
abbreviations = []
for material in insulation_types:
    abbreviation = ''.join(word[0].upper() for word in material.split())
    abbreviations.append(abbreviation)

# Create lists for metrics and values
metrics = ['Embodied Carbon - kgCO2e/kg','DQI Score']
x = np.arange(len(insulation_df['Materials']))
width = 0.15

# Create a figure
fig, ax = plt.subplots(figsize=(30,15))

# Loop through each metric
for i, metric in enumerate(metrics):
    ax.bar(x + i * width, insulation_df[metric], width, label=metric)

ax.set_xlabel('Insulation Type', fontsize=14)
ax.set_ylabel('Values', fontsize=14)
ax.set_title('Metrics vs Values for Insulation Material', fontsize=16)
ax.set_xticks(x + 0.5 * width * (len(metrics) - 1))
ax.set_xticklabels(abbreviations, rotation=45, ha='right', fontsize=12)
ax.legend(fontsize=12)

# Add a note with the abbreviations
note = '\n'.join(f'{abbreviation}: {material}' for abbreviation, material in zip(abbreviations, insulation_types))
ax.annotate(note, xy=(1.05, 0.5), xycoords='axes fraction', ha='left', fontsize=12)

# Layout so plots do not overlap
fig.tight_layout(rect=[0,0,0.75,1])
plt.tight_layout()
plt.savefig('plot8.png', bbox_inches='tight')
plt.show()

#material data sheet 2
#plotting 9
# Load the Excel file
df = pd.read_excel('FinalDataSet.xlsx', sheet_name='Material Data')

def extract_material_group(name):
    name = str(name).lower()
    if 'steel' in name:
        return 'Steel'
    elif 'concrete' in name:
        return 'Concrete'
    elif 'timber' in name:
        return 'Timber'
    elif 'glass' in name:
        return 'Glass'
    elif 'aluminum' in name or 'aluminium' in name:
        return 'Aluminum'
    elif 'asphalt' in name:
        return 'Asphalt'
    elif 'bitumen' in name:
        return 'Bitumen'
    elif 'cement' in name:
        return 'Cement'
    elif 'aggregate' in name:
        return 'Aggregate'
    elif 'clay' in name:
        return 'Clay'
    elif 'paint' in name:
        return 'Paint'
    elif 'vinyl' in name:
        return 'Vinyl'
    elif 'insulation' in name:
        return 'Insulation'
    elif 'rubber' in name:
        return 'Rubber'
    elif 'plaster' in name:
        return 'Plaster'
    else:
        return 'Other'

# Apply the function to the 'Material' column
df['Material Group'] = df['Main Material'].apply(extract_material_group)

# Convert 'Sample Size' column to numeric values
df['Sample Size'] = pd.DataFrame(pd.to_numeric(df['Sample Size'], errors='coerce'))

# Drop rows with non-numeric values
df = df.dropna(subset=['Sample Size'])

# Group by material group and calculate sum of sample size
grouped_df = df.groupby('Material Group')['Sample Size'].sum().reset_index()

# Create a bar chart of material group vs sample size
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['Sample Size'])
plt.xlabel('Material Group')
plt.ylabel('Sample Size')
plt.title('Material Group vs Sample Size')
plt.xticks(rotation=90)
plt.tight_layout()
plt.savefig('plot_9.png', bbox_inches='tight')
plt.show()

#plot 10
N = 10
top_materials = df.groupby('Main Material')['Sample Size'].sum().sort_values(ascending=False).head(N)

# Create a bar chart of main material vs sample size
plt.figure(figsize=(10, 6))
plt.bar(top_materials.index, top_materials.values)
plt.xlabel('Main Material')
plt.ylabel('Sample Size')
plt.title('Top {} Main Materials by Sample Size'.format(N))
plt.xticks(rotation=90)
plt.tight_layout()
plt.tight_layout()
plt.savefig('plot_10.png', bbox_inches='tight')
plt.show()

#plot 11

# Apply the function to the 'Material' column
df['Material Group'] = df['Main Material'].apply(extract_material_group)

# Convert 'Embodied Carbon - kgCO2e/kg' column to numeric values
df['Max Embodied Carbon (kg CO2e/kg)'] = pd.to_numeric(df['Max Embodied Carbon (kg CO2e/kg)'], errors='coerce')

# Drop rows with non-numeric values
df = df.dropna(subset=['Max Embodied Carbon (kg CO2e/kg)'])

# Group by material group and calculate max embodied carbon
grouped_df = df.groupby('Material Group')['Max Embodied Carbon (kg CO2e/kg)'].max().reset_index()

# Create a bar chart of material group vs max embodied carbon
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['Max Embodied Carbon (kg CO2e/kg)'])
plt.xlabel('Material Group')
plt.ylabel('Max Embodied Carbon (kg CO2e/kg)')
plt.title('Material Group vs Max Embodied Carbon')
plt.xticks(rotation=90)
plt.tight_layout()
plt.savefig('plot_11.png', bbox_inches='tight')
plt.show()

#plot 12

# Convert 'Embodied Carbon - kgCO2e/kg' column to numeric values
df['Standard Deviation Embodied Carbon (kg CO2e/kg)'] = pd.to_numeric(df['Standard Deviation Embodied Carbon (kg CO2e/kg)'], errors='coerce')

# Drop rows with non-numeric values
df = df.dropna(subset=['Standard Deviation Embodied Carbon (kg CO2e/kg)'])

# Group by material group and calculate standard deviation of embodied carbon
grouped_df = df.groupby('Material Group')['Standard Deviation Embodied Carbon (kg CO2e/kg)'].std().reset_index()

# Create a bar chart of material group vs standard deviation of embodied carbon
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['Standard Deviation Embodied Carbon (kg CO2e/kg)'])
plt.xlabel('Material Group')
plt.ylabel('Standard Deviation Embodied Carbon (kg CO2e/kg)')
plt.title('Material Group vs Standard Deviation Embodied Carbon')
plt.xticks(rotation=90)
plt.tight_layout()
plt.savefig('plot_12.png', bbox_inches='tight')
plt.show()

#plot 13
# Convert 'Embodied Carbon - kgCO2e/kg' column to numeric values
df['Variance Based on a Sample (Excel Var Function) Embodied Carbon (kg CO2e/kg)'] = pd.to_numeric(df['Variance Based on a Sample (Excel Var Function) Embodied Carbon (kg CO2e/kg)'], errors='coerce')

# Drop rows with non-numeric values
df = df.dropna(subset=['Variance Based on a Sample (Excel Var Function) Embodied Carbon (kg CO2e/kg)'])

# Group by material group and calculate variance of embodied carbon
grouped_df = df.groupby('Material Group')['Variance Based on a Sample (Excel Var Function) Embodied Carbon (kg CO2e/kg)'].var().reset_index()

# Create a bar chart of material group vs variance of embodied carbon
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['Variance Based on a Sample (Excel Var Function) Embodied Carbon (kg CO2e/kg)'])
plt.xlabel('Material Group')
plt.ylabel('Variance Embodied Carbon (kg CO2e/kg)')
plt.title('Material Group vs Variance Embodied Carbon')
plt.xticks(rotation=90)
plt.tight_layout()
plt.savefig('plot_13.png', bbox_inches='tight')
plt.show()


#plot 14
# Load the Excel file
df = pd.read_excel('FinalDataSet.xlsx', sheet_name='Material_Data1')

# Apply the function to the 'Main Material' column
df['Material Group'] = df['Main Material'].apply(extract_material_group)

# Group the data by material group and calculate the sum of sample size
grouped_df = df.groupby('Material Group')['Sample Size'].sum().reset_index()

# Create a bar chart of material groups vs sample size
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['Sample Size'])
plt.xlabel('Material Group')
plt.ylabel('Sample Size')
plt.title('Sample Size by Material Group')
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('plot_14.png', bbox_inches='tight')
plt.show()

#plot 15

# Apply the function to the 'Main Material' column
df['Material Group'] = df['Main Material'].apply(extract_material_group)

# Group the data by material group and calculate the sum of sample size
grouped_df = df.groupby('Material Group')['DQI Sample Size (Max 10)'].sum().reset_index()

# Create a bar chart of material groups vs sample size
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['DQI Sample Size (Max 10)'])
plt.xlabel('Material Group')
plt.ylabel('DQI Sample Size (Max 10)')
plt.title('Sample Size by Material Group')
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('plot_15.png', bbox_inches='tight')
plt.show()

#plot 16
# Apply the function to the 'Main Material' column
df['Material Group'] = df['Main Material'].apply(extract_material_group)

# Group the data by material group and calculate the sum of sample size
grouped_df = df.groupby('Material Group')['DQI Total - % (Max 100%)'].sum().reset_index()

# Create a bar chart of material groups vs sample size
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['DQI Total - % (Max 100%)'])
plt.xlabel('Material Group')
plt.ylabel('DQI Total - % (Max 100%)')
plt.title('Sample Size by Material Group')
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('plot_16.png', bbox_inches='tight')
plt.show()

#plot 17

# Apply the function to the 'Main Material' column
df['Material Group'] = df['Main Material'].apply(extract_material_group)

# Group the data by material group and calculate the sum of sample size
grouped_df = df.groupby('Material Group')['DQI Temporal (Max 5)'].sum().reset_index()

# Create a bar chart of material groups vs sample size
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['DQI Temporal (Max 5)'])
plt.xlabel('Material Group')
plt.ylabel('DQI Temporal (Max 5)')
plt.title('Sample Size by Material Group')
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('plot_17.png', bbox_inches='tight')
plt.show()

#plot 18

# Group the data by material group and calculate the sum of sample size
grouped_df = df.groupby('Material Group')['DQI Geographic (Max 5)'].sum().reset_index()

# Create a bar chart of material groups vs sample size
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['DQI Geographic (Max 5)'])
plt.xlabel('Material Group')
plt.ylabel('DQI Geographic (Max 5)')
plt.title('Sample Size by Material Group')
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('plot_18.png', bbox_inches='tight')
plt.show()

#plot 19


# Group the data by material group and calculate the sum of sample size
grouped_df = df.groupby('Material Group')['DQI Transparency (Max 5)'].sum().reset_index()

# Create a bar chart of material groups vs sample size
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['DQI Transparency (Max 5)'])
plt.xlabel('Material Group')
plt.ylabel('DQI Transparency (Max 5)')
plt.title('Sample Size by Material Group')
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('plot_19.png', bbox_inches='tight')
plt.show()

#plot 20

# Load the Excel file
df = pd.read_excel('FinalDataSet.xlsx', sheet_name='Material_Data2')

def extract_material_group(name):
    name = name.lower()
    if 'steel' in name:
        return 'Steel'
    elif 'concrete' in name:
        return 'Concrete'
    elif 'timber' in name:
        return 'Timber'
    elif 'glass' in name:
        return 'Glass'
    elif 'aluminum' in name or 'aluminium' in name:
        return 'Aluminum'
    elif 'asphalt' in name:
        return 'Asphalt'
    elif 'bitumen' in name:
        return 'Bitumen'
    elif 'cement' in name:
        return 'Cement'
    elif 'aggregate' in name:
        return 'Aggregate'
    elif 'clay' in name:
       return 'Clay'
    elif 'paint' in name:
        return 'Paint'
    elif 'vinyl' in name:
        return 'Vinyl'
    elif 'insulation' in name:
        return 'Insulation'
    elif 'rubber' in name:
        return 'Rubber'
    elif 'plaster' in name:
        return 'Plaster'
    else:
        return 'Other'

# Apply the function to the 'Material' column
df['Material Group'] = df['Material'].apply(extract_material_group)

# Convert 'Density (kg m-3)' column to numeric values
df['Density (kg m-3)'] = pd.to_numeric(df['Density (kg m-3)'], errors='coerce')

# Drop rows with non-numeric values
df = df.dropna(subset=['Density (kg m-3)'])

# Group by material group and calculate mean density
grouped_df = df.groupby('Material Group')['Density (kg m-3)'].mean().reset_index()

# Create a bar chart of material group vs density
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['Density (kg m-3)'])
plt.xlabel('Material Group')
plt.ylabel('Density (kg m-3)')
plt.title('Material Group vs Density')
plt.xticks(rotation=90)
plt.tight_layout()
plt.savefig('plot_20.png', bbox_inches='tight')
plt.show()

#plot 21
# Convert 'Specific heat (J kg-1 K-1)' column to numeric values
df['Specific heat (J kg-1 K-1)'] = pd.to_numeric(df['Specific heat (J kg-1 K-1)'], errors='coerce')

# Drop rows with non-numeric values
df = df.dropna(subset=['Specific heat (J kg-1 K-1)'])

def extract_material_group(name):
    name = name.lower()
    if 'steel' in name:
        return 'Steel'
    elif 'concrete' in name:
        return 'Concrete'
    elif 'timber' in name:
        return 'Timber'
    elif 'glass' in name:
        return 'Glass'
    elif 'aluminum' in name or 'aluminium' in name:
        return 'Aluminum'
    elif 'asphalt' in name:
        return 'Asphalt'
    elif 'bitumen' in name:
        return 'Bitumen'
    elif 'cement' in name:
        return 'Cement'
    elif 'aggregate' in name:
        return 'Aggregate'
    elif 'clay' in name:
        return 'Clay'
    elif 'paint' in name:
        return 'Paint'
    elif 'vinyl' in name:
        return 'Vinyl'
    elif 'insulation' in name:
        return 'Insulation'
    elif 'rubber' in name:
        return 'Rubber'
    elif 'plaster' in name:
        return 'Plaster'
    else:
        return 'Other'

# Apply the function to the 'Material' column
df['Material Group'] = df['Material'].apply(extract_material_group)

# Group by material group and calculate mean specific heat
grouped_df = df.groupby('Material Group')['Specific heat (J kg-1 K-1)'].mean().reset_index()

# Create a bar chart of material group vs specific heat
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['Specific heat (J kg-1 K-1)'])
plt.xlabel('Material Group')
plt.ylabel('Specific heat (J kg-1 K-1)')
plt.title('Material Group vs Specific heat')
plt.xticks(rotation=90)
plt.tight_layout()
plt.savefig('plot_21.png', bbox_inches='tight')
plt.show()

#plot 22
# Convert 'Thermal Diffusivity (M^2 S-1)' column to numeric values
df['Thermal Diffusivity (M^2 S-1)'] = pd.to_numeric(df['Thermal Diffusivity (M^2 S-1)'], errors='coerce')

# Drop rows with non-numeric values
df = df.dropna(subset=['Thermal Diffusivity (M^2 S-1)'])
def extract_material_group(name):
    name = name.lower()
    if 'steel' in name:
        return 'Steel'
    elif 'concrete' in name:
        return 'Concrete'
    elif 'timber' in name:
        return 'Timber'
    elif 'glass' in name:
        return 'Glass'
    elif 'aluminum' in name or 'aluminium' in name:
        return 'Aluminum'
    elif 'asphalt' in name:
        return 'Asphalt'
    elif 'bitumen' in name:
        return 'Bitumen'
    elif 'cement' in name:
        return 'Cement'
    elif 'aggregate' in name:
        return 'Aggregate'
    elif 'clay' in name:
        return 'Clay'
    elif 'paint' in name:
        return 'Paint'
    elif 'vinyl' in name:
        return 'Vinyl'
    elif 'insulation' in name:
        return 'Insulation'
    elif 'rubber' in name:
        return 'Rubber'
    elif 'plaster' in name:
        return 'Plaster'
    else:
        return 'Other'
# Apply the function to the 'Material' column
df['Material Group'] = df['Material'].apply(extract_material_group)

# Group by material group and calculate mean thermal diffusivity
grouped_df = df.groupby('Material Group')['Thermal Diffusivity (M^2 S-1)'].mean().reset_index()

# Create a bar chart of material group vs thermal diffusivity
plt.figure(figsize=(10, 6))
plt.bar(grouped_df['Material Group'], grouped_df['Thermal Diffusivity (M^2 S-1)'])
plt.xlabel('Material Group')
plt.ylabel('Thermal Diffusivity (M^2 S-1)')
plt.title('Material Group vs Thermal Diffusivity')
plt.xticks(rotation=90)
plt.tight_layout()
plt.savefig('plot_22.png', bbox_inches='tight')
plt.show()

#cost analysis 
# # Load the Material_Comparison_Dataset CSV file
material_df = pd.read_csv('Material_Comparison_Dataset.csv')

# Load the FinalDataSet Excel file
final_df = pd.read_excel('FinalDataSet.xlsx')

# Append the material data to the final data
final_df = pd.concat([final_df, material_df])

# Save the updated final data to the Excel file
final_df.to_excel('FinalDataSet.xlsx', index=False)

# Create an Excel writer
with pd.ExcelWriter('FinalDataSet.xlsx', mode='a', if_sheet_exists='replace') as writer:
    material_df.to_excel(writer, sheet_name='Material Cost',index=False)

# Read the Excel file
df = pd.read_excel('FinalDataSet.xlsx', sheet_name='Material Cost')

# Define the categories
categories = ['Cost per Unit (USD)', 'Recyclability Index', 'Embodied Carbon (kgCO?)', 'Energy Consumption (kWh)', 'Lifespan (Years)']

# Normalize the values
for category in categories:
    df[category] = df[category] / df[category].max()

# Plot the radar chart
fig = plt.figure(figsize=(10, 8))
ax = fig.add_subplot(111, polar=True)

for index, row in df.iterrows():
    values = [row[category] for category in categories]
    angles = np.linspace(0, 2*np.pi, len(categories), endpoint=False)
    values.append(values[0])
    angles = np.append(angles, angles[0])
    ax.plot(angles, values, 'o-', linewidth=2, label=row['Material'])

ax.set_thetagrids(angles[:-1] * 180/np.pi, categories)
ax.set_ylim(0, 1)
ax.set_title('Material Characteristics')
ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1))

plt.tight_layout()
plt.savefig('plot_23.png', bbox_inches='tight')
plt.show()